import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Excel Filter & Export Tool", layout="wide", initial_sidebar_state="expanded")

st.title("📊 Excel Filter & Export Tool")
st.caption("Upload any sheet → map columns → filter → export ward sheets + summary pivot")

EXCEL_ILLEGAL = r'\/:*?[]'

def sanitize_sheet_name(name: str, maxlen: int = 31) -> str:
    return "".join(c for c in str(name) if c not in EXCEL_ILLEGAL)[:maxlen] or "Sheet"

def best_match(columns: list, keywords: list, fallback_index: int = 0) -> int:
    for kw in keywords:
        for i, c in enumerate(columns):
            if kw in c.lower():
                return i
    return min(fallback_index, len(columns) - 1)

@st.cache_data(show_spinner="Reading file…")
def load_file(file_bytes: bytes, name: str) -> pd.DataFrame:
    try:
        if name.endswith(".csv"):
            return pd.read_csv(io.BytesIO(file_bytes), dtype=str, encoding="utf-8-sig")
        return pd.read_excel(io.BytesIO(file_bytes), dtype=str)
    except Exception as e:
        st.error(f"Could not read file: {e}")
        st.stop()

# ── Upload ────────────────────────────────────────────────────────────────────
uploaded = st.file_uploader("Upload Excel (.xlsx / .xls) or CSV", type=["xlsx", "xls", "csv"])
if not uploaded:
    st.info("⬆️ Upload a file to get started.")
    st.stop()

raw_bytes = uploaded.read()
df_raw = load_file(raw_bytes, uploaded.name)
df_raw.columns = df_raw.columns.str.strip()
all_cols = df_raw.columns.tolist()

st.success(f"✅ Loaded **{len(df_raw):,}** rows × **{len(all_cols)}** columns from `{uploaded.name}`")

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ Column Mapping")
    st.caption("Auto-detected — override if needed")

    zone_idx = best_match(all_cols, ["zonename", "zone_name", "zone"])
    ward_idx = best_match(all_cols, ["wardname", "ward_name", "ward"])

    zone_col = st.selectbox("🗺️ Zone column",  all_cols, index=zone_idx)
    ward_col = st.selectbox("🏘️ Ward column",  all_cols, index=ward_idx)

    st.divider()
    st.subheader("📐 Pivot columns")
    st.caption("Columns to SUM in the pivot table")

    numeric_hints = ["amt", "amount", "total", "paid", "fee", "charge",
                     "vat", "income", "book", "corp", "form", "other", "govt"]
    auto_numeric = [c for c in all_cols if any(k in c.lower() for k in numeric_hints)]

    pivot_cols = st.multiselect("Sum columns", options=all_cols,
                                default=auto_numeric if auto_numeric else [])

    st.divider()
    st.subheader("🔍 Zone Filter")
    all_zones = sorted(df_raw[zone_col].dropna().unique().tolist())
    sel_zones = st.multiselect("Include zones", all_zones, default=all_zones)

    st.divider()
    st.subheader("🧹 Cleaning")
    remove_dash = st.checkbox("Strip trailing dash ( - ) from column", value=False)
    if remove_dash:
        dash_col_idx = best_match(all_cols, ["holding", "e-holding", "doc", "id"])
        dash_col = st.selectbox("Column to clean", all_cols, index=dash_col_idx)
    else:
        dash_col = None

    st.divider()
    st.subheader("👁️ Column Visibility")
    st.caption("Uncheck to exclude from export")
    col_flags = {c: st.checkbox(c, value=True, key=f"vis_{c}") for c in all_cols}

# ── Transforms ────────────────────────────────────────────────────────────────
df = df_raw.copy()
if sel_zones:
    df = df[df[zone_col].isin(sel_zones)]
df = df[[c for c, v in col_flags.items() if v]]
if remove_dash and dash_col and dash_col in df.columns:
    df[dash_col] = df[dash_col].astype(str).str.rstrip("-").str.strip()

# ── Preview ───────────────────────────────────────────────────────────────────
st.subheader("📋 Preview (first 300 rows)")
st.dataframe(df.head(300), use_container_width=True, height=320)
st.caption(f"Filtered: **{len(df):,}** rows | **{len(df.columns)}** columns")

# ── Excel Builder ─────────────────────────────────────────────────────────────
def build_excel(df: pd.DataFrame, zone_col: str, ward_col: str, pivot_cols: list) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    HDR_FILL   = PatternFill("solid", start_color="1F4E79")
    HDR_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    DATA_FONT  = Font(name="Arial", size=10)
    ALT_FILL   = PatternFill("solid", start_color="D6E4F0")
    TOTAL_FILL = PatternFill("solid", start_color="C6EFCE")
    BOLD_FONT  = Font(bold=True, name="Arial", size=10)
    CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT       = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    RIGHT      = Alignment(horizontal="right",  vertical="center")
    thin       = Side(style="thin", color="A0A0A0")
    BORDER     = Border(left=thin, right=thin, top=thin, bottom=thin)

    def apply_header(ws, ncols):
        for ci in range(1, ncols + 1):
            cell = ws.cell(row=1, column=ci)
            cell.fill = HDR_FILL; cell.font = HDR_FONT
            cell.alignment = CENTER; cell.border = BORDER
        ws.row_dimensions[1].height = 20

    def auto_width(ws):
        for col in ws.columns:
            w = max((len(str(c.value)) if c.value is not None else 0 for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 3, 42)

    def write_sheet(ws, data: pd.DataFrame, num_cols: list):
        cols = data.columns.tolist()
        for ci, col in enumerate(cols, 1):
            ws.cell(row=1, column=ci, value=col)
        apply_header(ws, len(cols))

        for ri, (_, row) in enumerate(data.iterrows(), 2):
            is_alt = ri % 2 == 0
            for ci, (col, val) in enumerate(zip(cols, row), 1):
                if col in num_cols:
                    try: val = float(val)
                    except (ValueError, TypeError): pass
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = DATA_FONT; cell.border = BORDER
                if col in num_cols and isinstance(val, float):
                    cell.alignment = RIGHT; cell.number_format = '#,##0.##'
                elif ci == 1:
                    cell.alignment = LEFT
                else:
                    cell.alignment = CENTER
                if is_alt:
                    cell.fill = ALT_FILL

        auto_width(ws)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    # Ward sheets
    valid_pivot = [c for c in pivot_cols if c in df.columns]
    summary_rows = []
    seen = {}

    for zone_val in sorted(df[zone_col].dropna().unique()):
        zone_df = df[df[zone_col] == zone_val]
        for ward_val in sorted(zone_df[ward_col].dropna().unique()):
            ward_df = zone_df[zone_df[ward_col] == ward_val].reset_index(drop=True)

            raw_name = sanitize_sheet_name(f"{zone_val}_{ward_val}")
            if raw_name in seen:
                seen[raw_name] += 1
                sheet_name = sanitize_sheet_name(raw_name[:28] + f"_{seen[raw_name]}")
            else:
                seen[raw_name] = 0
                sheet_name = raw_name

            ws = wb.create_sheet(title=sheet_name)
            write_sheet(ws, ward_df, valid_pivot)

            row = {"Zone": zone_val, "Ward": ward_val, "Rows": len(ward_df)}
            for pc in valid_pivot:
                row[f"Sum_{pc}"] = pd.to_numeric(ward_df[pc], errors="coerce").sum()
            summary_rows.append(row)

    # Summary sheet
    ws_s = wb.create_sheet(title="Summary")
    summary_df = pd.DataFrame(summary_rows)

    if summary_df.empty:
        ws_s["A1"] = "No data matched the selected filters."
        wb.move_sheet("Summary", offset=-len(wb.sheetnames) + 1)
        out = io.BytesIO(); wb.save(out); return out.getvalue()

    sum_cols = [c for c in summary_df.columns if c.startswith("Sum_")]
    write_sheet(ws_s, summary_df, sum_cols)

    # Pivot block
    pr = len(summary_df) + 4
    title_cell = ws_s.cell(row=pr, column=1, value="PIVOT — Zone Totals")
    title_cell.font = Font(bold=True, size=12, name="Arial", color="1F4E79")

    pivot_headers = ["Zone", "Wards", "Total Rows"] + [f"Total {c[4:]}" for c in sum_cols]
    for ci, h in enumerate(pivot_headers, 1):
        cell = ws_s.cell(row=pr + 1, column=ci, value=h)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = CENTER; cell.border = BORDER

    agg = {"Wards": ("Ward", "nunique"), "Total Rows": ("Rows", "sum")}
    for sc in sum_cols:
        agg[f"Total {sc[4:]}"] = (sc, "sum")
    pivot_df = summary_df.groupby("Zone", as_index=False).agg(**agg)
    pivot_num = [f"Total {sc[4:]}" for sc in sum_cols]

    for ri, (_, row) in enumerate(pivot_df.iterrows(), pr + 2):
        is_alt = ri % 2 == 0
        for ci, (col, val) in enumerate(zip(pivot_headers, row), 1):
            if col in pivot_num:
                try: val = float(val)
                except (ValueError, TypeError): pass
            cell = ws_s.cell(row=ri, column=ci, value=val)
            cell.font = DATA_FONT; cell.border = BORDER
            if col in pivot_num and isinstance(val, float):
                cell.alignment = RIGHT; cell.number_format = '#,##0.##'
            elif ci == 1:
                cell.alignment = LEFT
            else:
                cell.alignment = CENTER
            if is_alt:
                cell.fill = ALT_FILL

    # Grand total
    gt = pr + 2 + len(pivot_df)
    ws_s.cell(row=gt, column=1, value="GRAND TOTAL")
    ds, de = pr + 2, gt - 1
    for ci in range(2, len(pivot_headers) + 1):
        cl = get_column_letter(ci)
        cell = ws_s.cell(row=gt, column=ci, value=f"=SUM({cl}{ds}:{cl}{de})")
        cell.number_format = '#,##0.##'; cell.alignment = RIGHT
    for ci in range(1, len(pivot_headers) + 1):
        cell = ws_s.cell(row=gt, column=ci)
        cell.fill = TOTAL_FILL; cell.font = BOLD_FONT; cell.border = BORDER

    auto_width(ws_s)
    wb.move_sheet("Summary", offset=-len(wb.sheetnames) + 1)

    out = io.BytesIO(); wb.save(out); return out.getvalue()

# ── Export ────────────────────────────────────────────────────────────────────
st.divider()
c1, c2, c3, c4 = st.columns(4)
c1.metric("Zones selected", len(sel_zones))
c2.metric("Total rows", f"{len(df):,}")
c3.metric("Pivot columns", len([c for c in pivot_cols if c in df.columns]))
if not df.empty and zone_col in df.columns and ward_col in df.columns:
    c4.metric("Ward sheets", df.groupby([zone_col, ward_col]).ngroups)

st.divider()

if zone_col not in df.columns or ward_col not in df.columns:
    st.warning("⚠️ Zone or Ward column missing from filtered data — check mapping & visibility.")
elif df.empty:
    st.warning("⚠️ No rows match the current filters.")
else:
    if st.button("🚀 Generate & Download Excel", type="primary", use_container_width=True):
        with st.spinner("Building workbook…"):
            excel_bytes = build_excel(df, zone_col, ward_col, pivot_cols)
        n = df.groupby([zone_col, ward_col]).ngroups
        st.success(f"✅ Done! **{n}** ward sheet(s) + **1** summary + pivot.")
        st.download_button(
            label="⬇️ Download Excel",
            data=excel_bytes,
            file_name="filtered_export.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
